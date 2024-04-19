"""
generates a random episode
"""

import random
import openpyxl as op
from openpyxl import load_workbook

# Name of data file
FILEPATH = "files/episodes.xlsx"

# Constants for selecting crack episodes
INCL = 0
EXCL = 1
ONLY = 2


class Episode:
    """
        name: episode name
        season: number of the season
        abs_num: number of the episode in total
        rel_num: number of the episode in the season
        is_crack: is a crack episode or not
    """
    def __init__(self,
                 ep_name:str,
                 ep_season:int,
                 ep_absolute_number:int,
                 ep_relative_number:int,
                 ep_is_crack:bool):
        self.name = ep_name
        self.season = ep_season
        self.abs_num = ep_absolute_number
        self.rel_num = ep_relative_number
        self.is_crack = ep_is_crack

    def __str__(self):
        return (str(self.season) + " " + str(self.rel_num) + " " + self.name)
    
    def get_name(self):
        return self.name
    
    def get_season(self):
        return self.season
    
    def get_abs_num(self):
        return self.abs_num
    
    def get_rel_num(self):
        return self.rel_num
    
    def get_is_crack(self):
        return self.is_crack


def get_all_episodes(filename:str):
    """
    filename: path of data (Excel) file
    """
    # open excel file and sheet (first sheet)
    ep_workbook = load_workbook(filename)
    ep_sheet = ep_workbook[ep_workbook.sheetnames[0]]

    # get first and last row that has data (not including title)
    first_row = 46
    last_row = ep_sheet.max_row
    last_row = 83 #TESTING PURPOSES
    first_column = ep_sheet.min_column
    last_column = ep_sheet.max_column

    # make a set of all episodes from the data
    all_eps = set([])
    for row in ep_sheet.iter_rows(first_row, last_row):
        # get episode info from corresponding columns
        season = row[0].value
        rel_num = row[1].value
        name = row[2].value
        # is crack if the column contains 1
        is_crack = (row[3].value == 1)
        # use (row number) - 1 for absolute episode number
        abs_num = row[1].row - 1

        # make episode object
        new_ep = Episode(name, season, abs_num, rel_num, is_crack)

        # add to list - order not guaranteed
        all_eps.add(new_ep)

    return all_eps


def get_random_episode(all_eps:set,
                       seasons:list = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15],
                       crack:int = INCL):
    """
    seasons: list with integers of seasons to be included; default is all
    crack: 1 (include crack and regular) or 2 (exclude crack, only regular) or 3 (crack episodes only, no regular); default is 1
    """
    # add all episodes that to randomly choose from to a new list
    eps_to_choose_from = []

    for ep in all_eps:
        # default to including in the list
        to_choose_from = True

        # exclude from list based on parameters from the form
        if crack == EXCL and ep.get_is_crack():
            to_choose_from = False
        elif crack == ONLY and not ep.get_is_crack():
            to_choose_from = False
        elif ep.get_season() not in seasons:
            to_choose_from = False

        # add episode to list if needed
        if to_choose_from:
            eps_to_choose_from.append(ep)

    # return random episode
    return random.choice(eps_to_choose_from)