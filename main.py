"""
generator and other functions
"""

import random
import openpyxl
from openpyxl import load_workbook


# Constants for selecting crack episodes
INCL = 0
EXCL = 1
ONLY = 2


class Episode:
    """
        name: episode name
        season: number of the season
        abs_num: number of the episode in total
        rel_num: number of the episode in the season
        is_crack: is a crack episode or not
    """
    def __init__(self,
                 ep_name:str,
                 ep_season:int,
                 ep_absolute_number:int,
                 ep_relative_number:int,
                 ep_is_crack:bool):
        self.name = ep_name
        self.season = ep_season
        self.abs_num = ep_absolute_number
        self.rel_num = ep_relative_number
        self.is_crack = ep_is_crack

    def __str__(self):
        return (str(self.season) + " " + str(self.rel_num) + " " + self.name)
    
    def get_name(self):
        return self.name
    
    def get_season(self):
        return self.season
    
    def get_abs_num(self):
        return self.abs_num
    
    def get_rel_num(self):
        return self.rel_num
    
    def get_is_crack(self):
        return self.is_crack


def get_all_episodes(filename:str):
    """
    filename: path of data (Excel) file
    """
    # open excel file and sheet (first sheet)
    ep_workbook = load_workbook(filename)
    ep_sheet = ep_workbook[ep_workbook.sheetnames[0]]

    # get first and last row that has data (not including title)
    first_row = 2
    last_row = ep_sheet.max_row
    first_column = ep_sheet.min_column
    last_column = ep_sheet.max_column

    # make a set of all episodes from the data
    all_eps = set([])
    for row in ep_sheet.iter_rows(first_row, last_row):
        # get episode info from corresponding columns
        season = row[0].value
        rel_num = row[1].value
        name = row[2].value
        # is crack if the column contains 1
        is_crack = (row[3].value == 1)
        # use (row number) - 1 for absolute episode number
        abs_num = row[1].row - 1

        # make episode object
        new_ep = Episode(name, season, abs_num, rel_num, is_crack)

        # add to list - order not guaranteed
        all_eps.add(new_ep)

    # close file
    ep_workbook.close()

    return all_eps


def get_random_episode(all_eps:set,
                       seasons:list = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15],
                       crack:int = INCL):
    """
    seasons: list with integers of seasons to be included; default is all
    crack: 0 (include crack and regular) or 1 (exclude crack, only regular) or 2 (crack episodes only, no regular); default is 0
    """
    # add all episodes that fit request to a new list
    eps_to_choose_from = []

    for ep in all_eps:

        if ep.get_season() not in seasons:
            # only keep episodes in requested seasons
            continue
        elif crack == EXCL and ep.get_is_crack():
            # only keep episodes that fit the requested "crack" parameter
            continue
        elif crack == ONLY and not ep.get_is_crack():
            continue

        # add episode to list if needed
        eps_to_choose_from.append(ep)
        
    return random.choice(eps_to_choose_from)


def parse_data(seasons_data:str, crack_data:str):
    """ 
    data: whatever was entered in the form; separate seasons with comma; connected seasons with dash
    crack_data: whatever was selected in the form (0,1,2)
    """
    seasons_int = []

    # split all the separate seasons
    seasons_str = seasons_data.split(",")

    # go through and convert to integers
    for i in seasons_str:
        if "-" in i:
            # is a range, so split get start and end season
            start, end = i.split("-")
            start = int(start)
            end = int(end)

            # add all numbers in range to list
            seasons_int.extend([j for j in range(start, end+1)])
        else:
            # single season, so just add the integer
            seasons_int.append(int(i))

    return seasons_int, int(crack_data)